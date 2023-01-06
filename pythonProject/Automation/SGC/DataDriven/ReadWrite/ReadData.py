import openpyxl

path_of_file= "/DataDriven/ReadWrite/data.xlsx"  #path of file

workbook=openpyxl.load_workbook(path_of_file)#to load excel file

#sheet=workbook.get_sheet_by_name("sheetanme")

sheet=workbook.active #3 sheet name of file from where you want to extract data

rows=sheet.max_row
cols=sheet.max_column

print("no of rows are", rows)
print("no of column are", cols)

for r in range(1, rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value, end="  ")


    print()









